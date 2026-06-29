"""Post-processing functions for report content.

Extracted from graph/report.py. Pure functions with no LLM dependency.
Handles figure embedding, renumbering, paragraph wrapping, and caption building.
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any


# Supported image extensions for filtering
PIC_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}

# Regex patterns
REF_BODY_BOUNDARY = re.compile(r'(<div\s+class=[\'"]ref-title[\'"]\s*>)')
# Matches image tags with optional following caption paragraphs
IMAGE_PATTERN = re.compile(
    r"!\[[^\]]*\]\((?P<path>[^)]+)\)\s*\n*\s*(<p\s+align='center'>.*?</p>(?:\s*\n*\s*<p\s+align='center'>.*?</p>)?)"
)
# Simple pattern to detect ANY embedded image (for deduplication)
_ANY_IMAGE_PATTERN = re.compile(r"!\[[^\]]*\]\((?P<path>[^)]+)\)")


def get_cited_numbers(body_content: str) -> set[int]:
    """Extract citation numbers that are actually used in body content."""
    return {
        int(m.group(1))
        for m in re.finditer(r'<sup>\[(\d+)\]</sup>', body_content)
    }


def filter_references_by_citation(references_block: str, cited_numbers: set[int]) -> str:
    """Keep only reference entries that are actually cited in the body."""
    if not references_block or not cited_numbers:
        return ""

    ref_pattern = re.compile(r'(<p[^>]*>\s*\[(\d+)\].*?</p>)', re.DOTALL)
    kept = [
        match.group(1)
        for match in ref_pattern.finditer(references_block)
        if int(match.group(2)) in cited_numbers
    ]
    return "\n\n".join(kept)


def build_caption_html(caption_title: str, caption_body: str, caption: str, output_lang: str = "zh-CN") -> str:
    """Build caption HTML from title and body or fallback to caption."""
    title = caption_title.strip()
    body = caption_body.strip() if caption_body else ""

    # Normalize title format based on output language
    if output_lang.lower().startswith("zh") or "chinese" in output_lang.lower():
        title = re.sub(r"Figure\s*\d+\s*\.?", "图 ", title, flags=re.IGNORECASE)
    else:
        title = re.sub(r"图\s*\d*\s*\.?", "Figure ", title)

    if title and body:
        return f"<p align='center'>{title}</p>\n\n<p align='center'>{body}</p>"
    return f"<p align='center'>{caption.strip()}</p>" if caption else ""


def embed_figures_in_body(body_content: str, figure_items: list[dict[str, str]], output_lang: str, start_index: int = 1) -> str:
    """Embed figures that are missing from body content.

    Strategy:
    1. Find which figures are already embedded in body
    2. For missing figures, append them at the end before references
    3. Renumber all figures based on final appearance order

    Args:
        start_index: First figure number to use (default 1). Set to 2 when the
            template already contains a static 图1 (e.g., 技术简介 workflow figure).
    """
    if not body_content or not figure_items:
        return body_content

    # Find which figures are already embedded in body
    # Use simple pattern to detect ANY image tag (regardless of caption format)
    # Normalize by filename since the agent may use different relative path prefixes
    embedded_filenames = set()
    for match in _ANY_IMAGE_PATTERN.finditer(body_content):
        path = match.group("path").strip()
        if path:
            embedded_filenames.add(Path(path).name)

    # Find missing figures (only actual image files, not scripts)
    missing_items = [
        item for item in figure_items
        if str(item.get("image_md_path", "")).strip()
        and Path(str(item.get("image_md_path", "")).strip()).suffix.lower() in PIC_EXTS
        and Path(str(item.get("image_md_path", "")).strip()).name not in embedded_filenames
    ]

    if not missing_items:
        return renumber_figures(body_content, figure_items, output_lang, start_index=start_index)

    # Build figure blocks for missing items
    def build_figure_block(item: dict[str, str], index: int) -> str:
        image_md_path = str(item.get("image_md_path", "")).strip()
        caption_title = str(item.get("caption_title", "")).strip()
        caption_body = str(item.get("caption_body", "")).strip()
        caption = str(item.get("caption", "")).strip()

        def update_figure_num(text: str) -> str:
            text = re.sub(r"Figure\s+\d+", f"Figure {index}", text, flags=re.IGNORECASE)
            return re.sub(r"图\s*\d+", f"图 {index}", text)

        caption_title = update_figure_num(caption_title)
        caption_body = update_figure_num(caption_body)

        caption_html = build_caption_html(caption_title, caption_body, caption, output_lang)
        figure_format = f"图 {index}" if output_lang.lower().startswith("zh") or "chinese" in output_lang.lower() else f"Figure {index}"
        return f"![{figure_format}]({image_md_path})\n\n{caption_html}"

    # Find references section and insert missing figures before it
    ref_match = REF_BODY_BOUNDARY.search(body_content)
    if ref_match:
        insert_pos = ref_match.start()
        figure_blocks = []
        for i, item in enumerate(missing_items, start=start_index):
            figure_blocks.append(build_figure_block(item, i))
        figures_text = "\n\n".join(figure_blocks)
        body_content = body_content[:insert_pos] + "\n\n" + figures_text + "\n\n" + body_content[insert_pos:]
    else:
        for i, item in enumerate(missing_items, start=start_index):
            body_content += "\n\n" + build_figure_block(item, i)

    return renumber_figures(body_content, figure_items, output_lang, start_index=start_index)


# Pattern to extract figure number from image tag: ![图 N](path) or ![Figure N](path)
_IMAGE_TAG_PATTERN = re.compile(r"!\[(图|Figure)\s*(\d+)\]")


def renumber_figures(body_content: str, figure_items: list[dict[str, str]], output_lang: str = "zh-CN", start_index: int = 1) -> str:
    """Renumber figures based on appearance order and rebuild caption HTML.

    Handles inconsistent agent numbering (index.md-based, zero-based, etc.)
    by replacing ALL figure numbers with sequential numbers based on document order.

    Args:
        start_index: First figure number to assign (default 1). Use 2 when the
            template already contains a static 图1 in 技术简介.
    """
    if not body_content or not figure_items:
        return body_content

    # Map filename → item for flexible path matching
    filename_to_item = {}
    for item in figure_items:
        md_path = str(item.get("image_md_path", "")).strip()
        if md_path:
            filename_to_item[Path(md_path).name] = item

    # Build ordered list of image paths from body (only real image files)
    ordered_paths = []
    seen_paths = set()
    for match in IMAGE_PATTERN.finditer(body_content):
        path = match.group("path").strip()
        if not path or path in seen_paths:
            continue
        if Path(path).suffix.lower() not in PIC_EXTS:
            continue
        if Path(path).name not in filename_to_item:
            continue
        ordered_paths.append(path)
        seen_paths.add(path)

    # Build path → new sequential index
    path_to_new_index = {path: str(idx) for idx, path in enumerate(ordered_paths, start=start_index)}

    def _swap(match: re.Match[str]) -> str:
        path = match.group("path").strip()
        if path not in path_to_new_index:
            return match.group(0)

        new_index = path_to_new_index[path]
        item = filename_to_item.get(Path(path).name, {})

        caption_html = match.group(2) if match.lastindex and match.lastindex >= 2 else ""

        if caption_html:
            def update_caption_nums(text: str) -> str:
                text = re.sub(r"Figure\s+\d+", f"Figure {new_index}", text, flags=re.IGNORECASE)
                return re.sub(r"图\s*\d+", f"图 {new_index}", text)

            para_pattern = re.compile(r"(<p\s+align='center'>.*?</p>)", re.DOTALL)
            paragraphs = para_pattern.findall(caption_html)

            if len(paragraphs) >= 2:
                updated_p1 = update_caption_nums(paragraphs[0])
                updated_p2 = update_caption_nums(paragraphs[1])
                caption_html = f"{updated_p1}\n\n{updated_p2}"
            elif len(paragraphs) == 1:
                caption_html = update_caption_nums(paragraphs[0])
        else:
            caption_title = str(item.get("caption_title", "")).strip()
            caption_body = str(item.get("caption_body", "")).strip()

            def update_figure_num(text: str) -> str:
                text = re.sub(r"Figure\s+\d+", f"Figure {new_index}", text, flags=re.IGNORECASE)
                return re.sub(r"图\s*\d+", f"图 {new_index}", text)

            caption_title = update_figure_num(caption_title)
            caption_body = update_figure_num(caption_body)
            caption_html = build_caption_html(caption_title, caption_body, "", output_lang)

        figure_format = f"图 {new_index}" if output_lang.lower().startswith("zh") or "chinese" in output_lang.lower() else f"Figure {new_index}"
        correct_path = str(item.get("image_md_path", path))
        return f"![{figure_format}]({correct_path})\n\n{caption_html}"

    body_content = IMAGE_PATTERN.sub(_swap, body_content)

    # Agent now uses outline-consistent numbering, so inline references are correct.
    # The old _strip_body_figure_refs call is kept as a no-op for backwards compatibility.

    return body_content


def _strip_body_figure_refs(body_content: str) -> str:
    """Remove figure references from body text paragraphs.

    The agent may use inconsistent numbering (index.md numbers, zero-based, etc.)
    that doesn't match the renumbered figure tags. Since figures are embedded
    adjacent to their discussion paragraphs, explicit references are unnecessary
    and potentially confusing when numbers don't match.
    """
    if not body_content:
        return body_content

    # Extract caption blocks to protect them
    caption_blocks: list[str] = []
    def _extract_captions(m: re.Match[str]) -> str:
        placeholder = f"__CAPTION_BLOCK_{len(caption_blocks)}__"
        caption_blocks.append(m.group(0))
        return placeholder

    body = re.sub(
        r"<p\s+align='center'>.*?</p>(?:\s*<p\s+align='center'>.*?</p>)?",
        _extract_captions,
        body_content,
        flags=re.DOTALL,
    )

    # Remove figure references in various formats:
    # 1. Parenthetical: （图5）, （图 5）, (Figure 5)
    body = re.sub(r'[（(]\s*(图|Figure)\s*\d+\s*[）)]', '', body)
    # 2. Inline with verb: 图 1 展示了, 图 4 显示, 如图 5 所示, 图 6 展示了
    body = re.sub(r'(?<!\[)(图|Figure)\s+(\d+)\s*(?=[展示显示呈现揭示富集的])', '', body, flags=re.IGNORECASE)
    # 3. Inline without space: 图0左图, 图2展示
    body = re.sub(r'(?<!\[)(图|Figure)\s*(\d+)(?=[左右上下前后中展示显示所示呈现富集])', '', body, flags=re.IGNORECASE)

    # Clean up double spaces or trailing punctuation artifacts
    body = re.sub(r'  +', ' ', body)
    body = re.sub(r'，\s*，', '，', body)
    body = re.sub(r'。\s*。', '。', body)

    # Restore caption blocks
    for i, caption in enumerate(caption_blocks):
        body = body.replace(f"__CAPTION_BLOCK_{i}__", caption)

    return body


def wrap_body_paragraphs(body_content: str) -> str:
    """Wrap plain text paragraphs in <p style='text-indent:18.20pt'> tags."""
    if not body_content:
        return body_content

    blocks = re.split(r'\n\s*\n', body_content)
    result = []

    for block in blocks:
        stripped = block.strip()
        if not stripped:
            continue

        first_line = stripped.split('\n')[0].strip()

        if (first_line.startswith('#')
                or first_line.startswith('!')
                or first_line.startswith('<')
                or first_line.startswith('- ')
                or first_line.startswith('* ')
                or first_line.startswith('|')
                or first_line.startswith('```')):
            result.append(block)
        else:
            result.append(f"<p style='text-indent:18.20pt'>{stripped}</p>")

    return '\n\n'.join(result)


def build_template_fields(report_output: dict[str, Any], report_title: str = "") -> dict[str, Any]:
    """Build template fields dictionary from report output.

    Args:
        report_output: Dictionary with body_md and other fields.
        report_title: Title for the report cover page. If empty, extracted from report_output.
    """
    title = report_title or str(report_output.get("report_title", ""))
    return {
        "Report_Title": title,
        "Cover_Report_Title": title,
        "Cover_Copyright_Text": "©2026All Rights Reserved",
        "Body_Content": str(report_output.get("body_md", "")),
    }


# Mapping from table file name (in project's table/ or tables/ folder) to template placeholder.
# Both .xlsx (legacy) and .csv (new) naming conventions are supported; the loader
# picks whichever file exists for each placeholder.
_TABLE_FILE_CANDIDATES: dict[str, list[str]] = {
    "Table_Project_Info": [
        "table1_project_info.csv",
        "01_project_info.xlsx",
    ],
    "Table_QC": [
        "table2_sequencing_quality.csv",
        "02_qc_stats.xlsx",
    ],
    "Table_Mapping": [
        "table3_mapping_statistics.csv",
        "03_mapping.xlsx",
    ],
    "Table_Gene_Capture": [
        "table4_cell_metrics.csv",
        "04_gene_capture.xlsx",
    ],
}

# Per-table cell style. Project info table uses a borderless wider style; the
# three sequencing tables share a bordered compact style.
_CELL_STYLE_PROJECT_INFO = (
    "border:none; padding:4px 12px;"
)
_CELL_STYLE_DATA = (
    "border:1px solid #7c97c8; padding:4px 10px; text-align:left; "
    "white-space:normal; overflow-wrap:anywhere; word-break:break-word;"
)
_PAR_STYLE_PROJECT_INFO = "margin:0; font-size:10pt; line-height:1.25;"
_PAR_STYLE_DATA = "margin:0; font-size:9pt; line-height:1.05;"

_TABLE_CELL_STYLE = {
    "Table_Project_Info": (_CELL_STYLE_PROJECT_INFO, _PAR_STYLE_PROJECT_INFO),
    "Table_QC": (_CELL_STYLE_DATA, _PAR_STYLE_DATA),
    "Table_Mapping": (_CELL_STYLE_DATA, _PAR_STYLE_DATA),
    "Table_Gene_Capture": (_CELL_STYLE_DATA, _PAR_STYLE_DATA),
}


def _excel_to_row_html(xlsx_path: Path, placeholder: str) -> str:
    """Read an .xlsx file and generate <tr><td>...</td></tr> HTML for data rows.

    Row 1 is treated as a header row (skipped, since headers live in the
    template). Rows 2+ are emitted as table data rows.
    """
    import logging

    logger = logging.getLogger(__name__)

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl not installed; cannot read %s", xlsx_path)
        return ""

    cell_style, par_style = _TABLE_CELL_STYLE.get(placeholder, (_CELL_STYLE_DATA, _PAR_STYLE_DATA))

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows_html: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            continue  # skip header row
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip empty rows
        cells_html = []
        for val in row:
            text = "" if val is None else str(val)
            cells_html.append(
                f"<td style='{cell_style}'><p style='{par_style}'>{text}</p></td>"
            )
        rows_html.append("<tr>" + "".join(cells_html) + "</tr>")
    wb.close()
    return "\n".join(rows_html)


def _csv_to_row_html(csv_path: Path, placeholder: str) -> str:
    """Read a .csv file and generate <tr><td>...</td></tr> HTML for data rows.

    Row 1 is treated as a header row (skipped, since headers live in the
    template). Rows 2+ are emitted as table data rows.
    """
    import csv

    cell_style, par_style = _TABLE_CELL_STYLE.get(placeholder, (_CELL_STYLE_DATA, _PAR_STYLE_DATA))

    rows_html: list[str] = []
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader):
            if row_idx == 0:
                continue  # skip header row
            if all(cell.strip() == "" for cell in row):
                continue  # skip empty rows
            cells_html = [
                f"<td style='{cell_style}'><p style='{par_style}'>{val.strip()}</p></td>"
                for val in row
            ]
            rows_html.append("<tr>" + "".join(cells_html) + "</tr>")
    return "\n".join(rows_html)


def load_table_files(project_path: str) -> dict[str, str]:
    """Load table data from <project_path>/{table,tables}/ and generate row HTML.

    For each placeholder, the loader tries known file names in order (CSV first,
    then XLSX) inside either `table/` or `tables/`. The first match wins.

    Row 1 of each file is a header (skipped — headers live in the template).
    Remaining rows are converted to <tr><td>...</td></tr> HTML. The template
    wraps these rows with its own <table>/<thead>/<tbody> and field descriptions.

    Returns:
        Dict mapping placeholder name (e.g., "Table_QC") to row HTML string.
        Missing files map to empty string with a WARNING log.
    """
    import logging

    logger = logging.getLogger(__name__)

    root = Path(project_path).expanduser().resolve()
    # Support both `table/` (legacy) and `tables/` (new convention).
    table_dirs = [root / "table", root / "tables"]

    result: dict[str, str] = {}
    for placeholder, candidates in _TABLE_FILE_CANDIDATES.items():
        found = None
        for tdir in table_dirs:
            for fname in candidates:
                fp = tdir / fname
                if fp.exists():
                    found = fp
                    break
            if found:
                break

        if not found:
            logger.warning("Table file not found for %s (searched %s)", placeholder, table_dirs)
            result[placeholder] = ""
            continue

        if found.suffix.lower() == ".csv":
            result[placeholder] = _csv_to_row_html(found, placeholder)
        else:
            result[placeholder] = _excel_to_row_html(found, placeholder)
        logger.info("Loaded table %s (%d rows)", found.name, result[placeholder].count("<tr>"))
    return result
